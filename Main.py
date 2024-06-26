import openpyxl as xl
import numpy as np

import math

FILE_NAME = 'book_output.xlsx'
# считываю данные из эксель файла бук_инпут
book_input = xl.open("book_input.xlsx")
sheet_input = book_input.active
a = sheet_input["B1"].value
ea = sheet_input["B11"].value
ei = sheet_input["B12"].value

# Считываю файл бук_аутпут, если таковой нет то создаю новую
try:
    book_output = xl.load_workbook(FILE_NAME)
except:
    book_output = xl.Workbook()

# удаляю дефолтный лист
for sheet_name in book_output.sheetnames:
    sheet = book_output[sheet_name]
    book_output.remove(sheet)

# получаю матрицу КоР из файла эксель
i = 1
KoR = np.array([])
while sheet_input[9][i].value != None:
    KoR = np.append(arr=KoR, values=sheet_input[9][i].value)
    i += 1
KoR = np.array([KoR])

# получаю матрицу Вектора нагружения из файла эксель
Rf = np.array([])
j = 2
while sheet_input[j][1].value != None:
    Rf = np.append(arr=Rf, values=sheet_input[j][1].value)
    j += 1

# print('Вектор нагружения=', '\n', Rf)
# print("Матрица KoR = ", "\n", KoR)

KoR = np.array([[0, 0, 0, 2]])


# Матрица направляющих косинусов
def DcosFramework(kor) -> np.ndarray:
    n, m = np.shape(kor)
    Le = np.zeros((n, 3))
    for i in range(0, n):
        Le[i][0] = math.sqrt((kor[i][2] - kor[i][0]) ** 2 + (kor[i][3] - kor[i][1]) ** 2)
        Le[i][1] = (kor[i][2] - kor[i][0]) / Le[i][0]
        Le[i][2] = (kor[i][3] - kor[i][1]) / Le[i][0]

    return Le




# Создается матрица Д. Сначала создаются 4 маленькие матрицы(r, r_dilda, delta, delta_tilda) а потом через np.concatenate соединяются друг с другом
def d_matrix_type0(dcos_matrix, EA) -> np.ndarray:
    r = np.zeros((4, 4))

    delta_tilda = np.array([[1 - (dcos_matrix[0][1] ** 2), dcos_matrix[0][1],
                             -1 - (dcos_matrix[0][1] ** 2), -dcos_matrix[0][1]]])
    r_tilda = -(np.transpose(delta_tilda))
    delta = np.array([[dcos_matrix[0][0] / (EA)]])

    sub1 = np.concatenate((r, delta_tilda), axis=0)
    sub2 = np.concatenate((r_tilda, delta), axis=0)
    d = np.concatenate((sub1, sub2), axis=1)

    return d


# Создается матрица Д. Сначала создаются 4 маленькие матрицы(r, r_dilda, delta, delta_tilda) а потом через np.concatenate соединяются друг с другом
def d_matrix_type1(dcos_matrix, EA, EI) -> np.ndarray:
    r = np.zeros((6, 6))

    L = dcos_matrix[0][0]
    cosa = dcos_matrix[0][1]
    sina = dcos_matrix[0][2]

    delta = np.zeros((3, 3))
    delta[0][0] = (((L) ** 3) / (12 * EI))
    delta[1][1] = (L / (EI))
    delta[2][2] = (L / (EA))

    r_tilda = np.array([[cosa, 0, -sina],
                        [-(L / 2), 1, 0],
                        [-sina, 0, -cosa],
                        [-cosa, 0, sina],
                        [-(L / 2), -1, 0],
                        [sina, 0, cosa]])

    delta_tilda = -(np.transpose(r_tilda))

    sub1 = np.concatenate((r, delta_tilda), axis=0)
    sub2 = np.concatenate((r_tilda, delta), axis=0)
    d = np.concatenate((sub1, sub2), axis=1)

    return d


# Создается матрица Д. Сначала создаются 4 маленькие матрицы(r, r_dilda, delta, delta_tilda) а потом через np.concatenate соединяются друг с другом
def d_matrix_type2(dcos_matrix, EA, EI) -> np.ndarray:
    r = np.zeros((6, 6))

    L = dcos_matrix[0][0]
    cosa = dcos_matrix[0][1]
    sina = dcos_matrix[0][2]

    delta = np.zeros((3, 3))
    delta[0][0] = (L / EI)
    delta[1][1] = (L / (3 * EI))
    delta[2][2] = (L / (3 * EA))
    delta[2][1] = (L / (6 * EI))
    delta[1][2] = (L / (6 * EI))

    delta_tilda = np.array([[cosa, sina, 0, -cosa, -sina, 0],
                            [sina / L, -(cosa / L), -1, -(sina / L), cosa / L, 0],
                            [-(sina / L), cosa / L, 0, sina / L, -(cosa / L), 1]])

    r_tilda = -(np.transpose(delta_tilda))

    sub1 = np.concatenate((r, delta_tilda), axis=0)
    sub2 = np.concatenate((r_tilda, delta), axis=0)
    d = np.concatenate((sub1, sub2), axis=1)

    return d


# удаляю ненужные столбцы
def remove_reactions(d_matrix) -> np.ndarray:
    i = 0
    while i < 3:
        d_matrix = np.delete(d_matrix, 0, 0)
        d_matrix = np.delete(d_matrix, 0, 1)
        i += 1

    return d_matrix


# Решаю матричное уравнение
def matrix_equation(d_with_opora) -> np.ndarray:
    d_with_opora = np.linalg.inv(d_with_opora)
    d_with_opora = d_with_opora.dot(-Rf)

    return d_with_opora


# функция перевода матриц из одного ряда в таблицы
def num_to_xl_special_hor(matrix, sheet_name):
    sheet_output = book_output.create_sheet(sheet_name)
    x = np.shape(matrix)[1]
    for i in range(x):
        sheet_output.cell(row=1, column=i + 1).value = matrix[0][i]


# функция перевода матриц из одной колонны в таблицы
def num_to_xl_special_vert(matrix, sheet_name):
    sheet_output = book_output.create_sheet(sheet_name)
    x = np.shape(matrix)[0]
    for i in range(x):
        sheet_output.cell(row=i + 1, column=1).value = matrix[i]


# функция перевода квадратных матриц в таблицы
def num_to_xl(matrix, sheet_name):
    sheet_output = book_output.create_sheet(sheet_name)
    x, y = np.shape(matrix)

    for i in range(x):
        for j in range(y):
            sheet_output.cell(row=j + 1, column=i + 1).value = matrix[j][i]


num_to_xl_special_hor(matrix=DcosFramework(KoR), sheet_name="COS")

num_to_xl_special_vert(matrix_equation(remove_reactions(d_matrix_type1(dcos_matrix=DcosFramework(KoR),
                                                                       EA=ea,
                                                                       EI=ei))), "matrix_equation_type1")
num_to_xl_special_vert(matrix_equation(remove_reactions(d_matrix_type2(dcos_matrix=DcosFramework(KoR),
                                                                       EA=ea,
                                                                       EI=ei))), "matrix_equation_type2")

num_to_xl(d_matrix_type0(dcos_matrix=DcosFramework(KoR),
                         EA=ea), "d_matrix_type0")

num_to_xl(d_matrix_type1(dcos_matrix=DcosFramework(KoR), EA=ea,
                         EI=ei), "d_matrix_type1")

num_to_xl(d_matrix_type2(dcos_matrix=DcosFramework(KoR), EA=ea,
                         EI=ei).round(3), "d_matrix_type2")

num_to_xl(remove_reactions(d_matrix_type1(dcos_matrix=DcosFramework(KoR),
                                          EA=ea,
                                          EI=ei)).round(3), "d_matrix2_reactions")

# Сохраняю и закрываю книгу
book_output.save(FILE_NAME)
book_output.close()

# print("Матрица d нулевого типа = ", "\n", d_matrix_type0(dcos_matrix=DcosFramework(KoR), EA=ea))

# print("матрица d первого типа = ", "\n", d_matrix_type1(dcos_matrix=DcosFramework(KoR),
# EA=ea,
# EI=ei))
# print("матрица d второго типа = ", "\n", d_matrix_type2(dcos_matrix=DcosFramework(KoR),
# EA=ea,
# EI=ei).round(3).astype(Decimal))

# print("Матрица d второго КЭ с жесткой звделкой = ", "\n",
# remove_reactions(d_matrix_type1(dcos_matrix=DcosFramework(KoR),
# EA=ea,
# EI=ei)).round(3).astype(Decimal))

# print("Решение матричного уравнение, матрица q, для КЭ первого типа = ", "\n",
# matrix_equation(remove_reactions(d_matrix_type1(dcos_matrix=DcosFramework(KoR),
# EA=ea,
# EI=ei))))

# print("Решение матричного уравнение, матрица q, для второго КЭ = ", "\n",
# matrix_equation(remove_reactions(d_matrix_type2(dcos_matrix=DcosFramework(KoR),
# EA=ea,
# EI=ei))).astype(Decimal))



